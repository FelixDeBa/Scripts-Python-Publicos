import openpyxl, re, subprocess
from datetime import date, timedelta
from openpyxl.styles import PatternFill
import os
from time import sleep
from tkinter import filedialog

os.system("")

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def get_last_fila(hoja, col):
    last_fila = 1
    while not hoja[f'{col}{last_fila}'].value is None:
        last_fila+=1
    return last_fila

def str_rango_semana(fecha):
    meses=('enero', 'febrero', 'marzo', 'abril', 'mayo','junio','julio','agosto','septiembre','agosto','octubre','noviembre','diciembre')
    lunes = fecha - timedelta(fecha.weekday())
    domingo = fecha + timedelta(6 - fecha.weekday()) 
    return f'{lunes.day} de {meses[lunes.month-1][:3]}-{domingo.day} de {meses[domingo.month-1][:3]}'


LINE_UP = '\033[1A'
LINE_CLEAR = '\x1b[2K'
hoy = date.today()
str_semana = str_rango_semana(hoy)

regexUsr = r'(?:Haciendo ping a |Pinging )(?P<dns>(?:.+)(?!\d{1,3}(?:\.\d{1,3}){3}))?(?: \[)?(?P<ip>\d{1,3}(?:\.\d{1,3}){3})(?:\])?'
regexIP = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
regexTimeout=r'(?:tiempo de espera agotado|request timed out)'
regexUnreach=r'(?:desintation host unreachable|host de destino inaccesible)'

ip_inicio = input("Desde que dirección IP comenzamos: ")
while not re.match(regexIP, ip_inicio):
    ip_inicio = input("Introduce una IP válida para iniciar: ")

octetos_inicio = [int(dato) for dato in ip_inicio.split('.')]

ip_fin = input("IP donde terminar de escanear (Por Defecto es 255.255.255.255): ")
if ip_fin == "" or ip_fin is None:
    ip_fin="255.255.255.255"
else:
    while not re.match(regexIP, ip_inicio):
        ip_fin = input("Introduce una IP válida: ")
        if ip_fin == "" or ip_fin is None:
            break

octetos_fin = [int(dato) for dato in ip_fin.split('.')]

pregunta = input("Desea crear un nuevo documento (Por defecto) o usar uno existente (E)? ")

if pregunta.lower() == "e":
    arch_ok=False
    while not arch_ok:
        nombredoc = filedialog.askopenfilename(title="Selecciona el archivo base")
        if nombredoc == "":
            print("No se seleccionó un archivo: ")
            arch_ok=False
        else:
            if (nombredoc.split('/')[-1]).split('.')[-1] in ['xlsx','xlsm','xls']:
                arch_ok=True
            else:
                arch_ok=False
                print('No seleccionaste un archivo compatible con excel')
    
    doc = openpyxl.load_workbook(nombredoc)
    hoja_enc = False
    for hoja_temp in doc.sheetnames:
        if [doc[hoja_temp][f'{chr(col+97)}1'].value for col in range(0,4)] == ['Fecha Escaneo', 'IP', 'Nombre Dominio', 'Informacion Adicional']:
            print(f'Trabajando sobre el archivo {nombredoc.split("/")[-1]} en la hoja {hoja_temp.title()}')
            hoja_enc = True
            hoja = doc[hoja_temp]
            fila = get_last_fila(doc[hoja_temp], 'a')
            break
    if not hoja_enc:
        hoja = doc.create_sheet('IPs' if not 'IPs' in doc.sheetnames else 'IPs_copia_')
        print(f'Trabajando sobre el archivo {nombredoc.split("/")[-1]} en la nueva hoja {hoja.title()}')
        hoja['a1'].value ='Fecha Escaneo'
        hoja['b1'].value ='IP'
        hoja['FW-MTYc1'].value ='Nombre Dominio'
        hoja['d1'].value ='Informacion Adicional'
        fila = 2
else:
    doc = openpyxl.Workbook()
    hoja = doc['Sheet']
    hoja.title = 'IPs'

    hoja['a1'].value ='Fecha Escaneo'
    hoja['b1'].value ='IP'
    hoja['c1'].value ='Nombre dominio'
    hoja['d1'].value ='Informacion Adicional'
    fila = 2
    print(f'Trabajando sobre nuevo archivo Escaneo_{str_semana}.xlsx en la hoja {hoja.title()}')
    nombredoc = f'Escaneo_{str_semana}.xlsx'


doclog=f'LogEscaneo {str(hoy)}.log'

for oct1 in range(octetos_inicio[0], octetos_fin[0]+1):
    for oct2 in range(octetos_inicio[1], octetos_fin[1]+1):
        for oct3 in range(octetos_inicio[2], octetos_fin[2]+1):
            for oct4 in range(octetos_inicio[3], octetos_fin[3]+1):
                log = open(doclog,'a')
                ip = f"{oct1}.{oct2}.{oct3}.{oct4}"
                proc = subprocess.Popen(f'ping -n 2 -w 3 -a {ip}', stdout=subprocess.PIPE, shell=True)
                (out, err) = proc.communicate()
                hostname="No encontrado"
                if err == None:
                    texto = str(out)[2:].split('\\r\\n')
                    for linea in texto:
                        if re.match(regexUsr, linea):
                            matches = re.search(regexUsr, linea)
                            hostname = matches.group('dns') if matches.group('dns') is not None else 'No encontrado'
                            break
                    hoja[f'a{fila}'].value = hoy
                    hoja[f'b{fila}'].value = ip
                    hoja[f'c{fila}'].value = hostname
                    if (re.match(regexTimeout, texto[2].lower()) and re.match(regexTimeout, texto[3].lower())) and hostname == 'No encontrado':
                        print(f'{bcolors.FAIL}{ip} no esta en Uso{bcolors.ENDC}')
                        log.writelines(f'\n{str(hoy)},{ip},{hostname},No está en uso')
                        print(LINE_UP, end=LINE_CLEAR)
                        hoja[f'd{fila}'].value = "No está en uso"
                        fila+=1
                    elif re.match(regexTimeout, texto[2].lower()) and re.match(regexTimeout, texto[3].lower()):
                        print(f'{bcolors.WARNING}IP: {ip}, usuario: {hostname} No respondió{bcolors.ENDC}')
                        log.writelines(f'\n{str(hoy)},{ip},{hostname},No respondió pero tiene asignado un hostname en Dominio')
                        print(LINE_UP, end=LINE_CLEAR)
                        hoja[f'd{fila}'].value = "Asignada pero no respondió"
                        fila+=1
                    elif re.match(regexUnreach, texto[2].lower()) and re.match(regexUnreach, texto[3].lower()):
                        print(f'{bcolors.WARNING}IP: {ip}, usuario: {hostname} Host inaccesible{bcolors.ENDC}')
                        log.writelines(f'\n{str(hoy)},{ip},{hostname},No respondió pero tiene asignado un hostname en Dominio')
                        print(LINE_UP, end=LINE_CLEAR)
                        hoja[f'd{fila}'].value = "Destino del host Inaccesible"
                        fila+=1
                    else:
                        print(f'{bcolors.OKBLUE}IP: {ip}, usuario: {hostname} OK{bcolors.ENDC}')
                        log.writelines(f'\n{str(hoy)},{ip},{hostname},OK')
                        print(LINE_UP, end=LINE_CLEAR)
                        hoja[f'd{fila}'].value = "OK"
                        fila+=1
                else:
                    print(f'{bcolors.FAIL}{bcolors.UNDERLINE}IP: {ip}, dió Error {err}{bcolors.ENDC}')
                    hoja[f'a{fila}'].value = hoy
                    hoja[f'a{fila}'].fill = PatternFill("solid", fgColor="FF0000")
                    hoja[f'b{fila}'].value = ip
                    hoja[f'b{fila}'].fill = PatternFill("solid", fgColor="FF0000")
                    hoja[f'c{fila}'].value = hostname
                    hoja[f'c{fila}'].fill = PatternFill("solid", fgColor="FF0000")
                    hoja[f'd{fila}'].value = err
                    hoja[f'd{fila}'].fill = PatternFill("solid", fgColor="FF0000")
                    fila+=1
                log.close()

sleep(2)
log = open(doclog,'a')
try:
    doc.save(nombredoc)
    print(f'Documento guardado como: {bcolors.OKCYAN}{bcolors.UNDERLINE}{nombredoc}{bcolors.ENDC}en la carpeta {bcolors.OKCYAN}{bcolors.UNDERLINE}{os.getcwd()}{bcolors.ENDC}')
except Exception as e:
    print(f'{bcolors.FAIL}{bcolors.UNDERLINE}No se pudo guardar el documento Error {e}{bcolors.ENDC}')
    log.writelines(f'{str(hoy)},No se pudo guardar el documento,Error {e},{str(e)}')
    
log.close()
print(f'Archivos de logs guardado como {doclog} en {os.getcwd()}')
os.startfile(os.getcwd())
